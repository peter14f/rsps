import io
import os.path
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (Font, Alignment, Border, Side,
                             PatternFill, Color, colors, fills)

ELEMENTARY = 0
MIDDLE = 1

requirement = [0, 3, 8, 9, 11, 13]

def get_size(start_path = '.'):
    total_size = 0
    for dirpath, dirnames, filenames in os.walk(start_path):
        for f in filenames:
            fp = os.path.join(dirpath, f)
            total_size += os.path.getsize(fp)
    return total_size

def missing_required_item(school_item):
    for i in requirement:
        if school_item[i] == '':
            return True
    return False

root_path = "Y:\\"
#root_path = "C:\\fakeZ"
print("Looking at directories in '" + root_path + "'")

root_list = os.listdir(root_path)
root_list = sorted(root_list)

#csvfile = open('summary.csv', 'w', newline='')
#csvwriter = csv.writer(csvfile)

wb = Workbook()
# grab the active worksheet
ws = wb.active

schools = []
items = []
elementary_school = []
elementary_school_done = []
middle_school = []
middle_school_done = []

for school in root_list:
    if re.match("^[0-9]", school):
        schools.append(school)

rsps_path = os.path.join(root_path, schools[21])
items_list = os.listdir(rsps_path)
cvt_map = {u'一': 10, u'二': 20, u'三': 30, u'四': 40}
items_map = {}

for item in items_list:
    value = int(item[4]) + cvt_map.get(item[1])
    items_map[value] = item

for key in sorted(items_map.keys()):
	items.append(items_map.get(key))

rows = []
first_row = ['', u'缺*成果']
col_offset = len(first_row)
first_row.extend(items)
rows.append(first_row)
p = re.compile("([0-9]+).+")


for school in schools:
    match = p.search(school)
    school_code = int(match.group(1))

    if school_code >= 300:
        middle_school.append(school)
        school_type = MIDDLE
    else:
        elementary_school.append(school)
        school_type = ELEMENTARY
    
    school_path = os.path.join(root_path, school)
    new_row = [school, '']
    school_item = []
    for item in items:
        item_path = os.path.join(school_path, item)
        size = get_size(item_path)
        if size > 0:
            school_item.append('V')
        else:
            school_item.append('')
    missing_requirement = missing_required_item(school_item)
    new_row.extend(school_item)
    if missing_requirement:
        new_row[1] = 'V'
    else:
        if school_type == ELEMENTARY:
            elementary_school_done.append(school)
        else:
            middle_school_done.append(school)
    rows.append(new_row)

for row in rows:
    ws.append(row)

for i in range(1, len(rows)+1):
    cell = "B%d" % i
    c = ws[cell]
    c.font = Font(bold=True)
    if i > 1:
        c.alignment= Alignment(horizontal='center')

# bottom
last_row = len(rows)
for j in range(65, 65+len(rows[0])-1):
    cell = "%c%d" % (j, last_row)
    c = ws[cell]
    c.border = Border(bottom=Side(style='thick'))

# right
last_col = 65+len(rows[0])-1
for i in range(1, len(rows)+1):
    cell = "%c%d" % (last_col, i)
    c = ws[cell]
    if i == len(rows):
        c.border = Border(right=Side(style='thick'),
                          bottom=Side(style='thick'))
    else:
        c.border = Border(right=Side(style='thick'))
    
ws.column_dimensions["%c" % last_col].width = 28

# divide
second_col = 66
for i in range(1, len(rows)+1):
    cell = "%c%d" % (second_col, i)
    c = ws[cell]
    if i == len(rows):
        c.border = Border(right=Side(style='thick'),
                          bottom=Side(style='thick'))
    else:
        c.border = Border(right=Side(style='thick'))

# center text
for j in range(67, 65+len(rows[0])):
    for i in range(2, len(rows)+1):
        cell = "%c%d" % (j, i)
        c = ws[cell]
        c.alignment = Alignment(horizontal='center')
        if j-col_offset-65 in requirement and c.value == "":
            c.fill = PatternFill(
                fill_type=fills.FILL_SOLID,
                start_color='0090EE90',
                end_color='0090EE90'
            )
# timestamp
now = datetime.now()

# Save the file
wb.save("summary_%s.xlsx" % now.strftime("%m-%d_%H_%M"))

#print(items)
#print (schools)

line1 = u"{}/{} 國小 {}".format(len(elementary_school_done), len(elementary_school), len(elementary_school) - len(elementary_school_done))
line2 = u"{}/{} 國中 {}".format(len(middle_school_done), len(middle_school), len(middle_school) - len(middle_school_done))
print(line1)
print(line2)
